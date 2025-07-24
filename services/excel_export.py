import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from flask import make_response

class ExcelExportService:
    """Service for exporting data to Excel format with professional styling"""
    
    @staticmethod
    def create_styled_workbook():
        """Create a workbook with professional styling"""
        wb = Workbook()
        return wb
    
    @staticmethod
    def style_header_row(ws, row_num, columns):
        """Apply professional styling to header row"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, column in enumerate(columns, 1):
            cell = ws.cell(row=row_num, column=col_num, value=column)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = 15
    
    @staticmethod
    def style_data_rows(ws, start_row, end_row, num_columns):
        """Apply styling to data rows"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, end_row + 1):
            for col in range(1, num_columns + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Alternate row colors
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    @staticmethod
    def export_purchase_orders():
        """Export Purchase Orders with items to Excel"""
        from models import PurchaseOrder, PurchaseOrderItem, Item, Supplier
        from app import db
        
        wb = ExcelExportService.create_styled_workbook()
        ws = wb.active
        ws.title = "Purchase Orders"
        
        # Add title
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = f"Purchase Orders Export - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Headers
        headers = [
            "PO Number", "Supplier", "PO Date", "Delivery Date", 
            "Status", "Item Code", "Item Name", "Quantity", "UOM", 
            "Rate", "Amount", "GST Rate", "Total Value"
        ]
        
        ExcelExportService.style_header_row(ws, 3, headers)
        
        # Data
        pos = db.session.query(PurchaseOrder).join(Supplier).all()
        row_num = 4
        
        for po in pos:
            po_items = db.session.query(PurchaseOrderItem).filter_by(purchase_order_id=po.id).all()
            
            if not po_items:
                # PO without items
                data = [
                    po.po_number, po.supplier.name, po.order_date.strftime('%Y-%m-%d'),
                    po.expected_date.strftime('%Y-%m-%d') if po.expected_date else '',
                    po.status.title(), '', '', '', '', '', '', '', ''
                ]
                for col_num, value in enumerate(data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
                row_num += 1
            else:
                # PO with items
                for item in po_items:
                    item_obj = db.session.query(Item).get(item.item_id)
                    
                    data = [
                        po.po_number, po.supplier.name, po.order_date.strftime('%Y-%m-%d'),
                        po.expected_date.strftime('%Y-%m-%d') if po.expected_date else '',
                        po.status.title(), 
                        item_obj.code if item_obj else '',
                        item_obj.name if item_obj else '',
                        item.qty, item.uom, item.rate, item.amount,
                        f"{item.gst_rate}%", item.total_price
                    ]
                    for col_num, value in enumerate(data, 1):
                        ws.cell(row=row_num, column=col_num, value=value)
                    row_num += 1
        
        # Style data rows
        ExcelExportService.style_data_rows(ws, 4, row_num - 1, len(headers))
        
        return wb
    
    @staticmethod
    def export_sales_orders():
        """Export Sales Orders with items to Excel"""
        from models import SalesOrder, SalesOrderItem, Item, Supplier
        from app import db
        
        wb = ExcelExportService.create_styled_workbook()
        ws = wb.active
        ws.title = "Sales Orders"
        
        # Add title
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = f"Sales Orders Export - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Headers
        headers = [
            "SO Number", "Customer", "Order Date", "Delivery Date", 
            "Status", "Item Code", "Item Name", "Quantity", "UOM", 
            "Rate", "Amount", "Total Value"
        ]
        
        ExcelExportService.style_header_row(ws, 3, headers)
        
        # Data
        sos = db.session.query(SalesOrder).join(Supplier).all()
        row_num = 4
        
        for so in sos:
            so_items = db.session.query(SalesOrderItem).filter_by(sales_order_id=so.id).all()
            
            if not so_items:
                # SO without items
                data = [
                    so.so_number, so.supplier.name, so.order_date.strftime('%Y-%m-%d'),
                    so.expected_date.strftime('%Y-%m-%d') if so.expected_date else '',
                    so.status.title(), '', '', '', '', '', '', ''
                ]
                for col_num, value in enumerate(data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
                row_num += 1
            else:
                # SO with items
                for item in so_items:
                    item_obj = db.session.query(Item).get(item.item_id)
                    
                    data = [
                        so.so_number, so.supplier.name, so.order_date.strftime('%Y-%m-%d'),
                        so.expected_date.strftime('%Y-%m-%d') if so.expected_date else '',
                        so.status.title(), 
                        item_obj.code if item_obj else '',
                        item_obj.name if item_obj else '',
                        item.quantity_ordered, item_obj.unit_of_measure if item_obj else '',
                        item.unit_price, item.total_price, item.total_price
                    ]
                    for col_num, value in enumerate(data, 1):
                        ws.cell(row=row_num, column=col_num, value=value)
                    row_num += 1
        
        # Style data rows
        ExcelExportService.style_data_rows(ws, 4, row_num - 1, len(headers))
        
        return wb
    
    @staticmethod
    def export_inventory():
        """Export Inventory data to Excel"""
        from models import Item
        from app import db
        
        wb = ExcelExportService.create_styled_workbook()
        ws = wb.active
        ws.title = "Inventory"
        
        # Add title
        ws.merge_cells('A1:K1')
        title_cell = ws['A1']
        title_cell.value = f"Inventory Export - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Headers
        headers = [
            "Item Code", "Item Name", "Description", "HSN Code", 
            "Current Stock", "UOM", "Purchase Unit", "Sale Unit",
            "Unit Price", "Business Type", "Min Stock Level"
        ]
        
        ExcelExportService.style_header_row(ws, 3, headers)
        
        # Data
        items = db.session.query(Item).all()
        row_num = 4
        
        for item in items:
            data = [
                item.code, item.name, item.description or '',
                item.hsn_code or '', item.current_stock, item.unit_of_measure,
                item.purchase_unit or '', item.sale_unit or '',
                item.unit_price, item.business_type or '', item.min_stock_level or 0
            ]
            for col_num, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
            row_num += 1
        
        # Style data rows
        ExcelExportService.style_data_rows(ws, 4, row_num - 1, len(headers))
        
        return wb
    
    @staticmethod
    def export_suppliers():
        """Export Suppliers data to Excel"""
        from models import Supplier
        from app import db
        
        wb = ExcelExportService.create_styled_workbook()
        ws = wb.active
        ws.title = "Suppliers"
        
        # Add title
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = f"Suppliers Export - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Headers
        headers = [
            "Name", "Partner Type", "Contact Person", "Mobile", 
            "Email", "GST Number", "Address", "City", "State"
        ]
        
        ExcelExportService.style_header_row(ws, 3, headers)
        
        # Data
        suppliers = db.session.query(Supplier).all()
        row_num = 4
        
        for supplier in suppliers:
            data = [
                supplier.name, supplier.partner_type or 'supplier',
                supplier.contact_person or '', supplier.mobile or '',
                supplier.email or '', supplier.gst_number or '',
                supplier.address or '', supplier.city or '', supplier.state or ''
            ]
            for col_num, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
            row_num += 1
        
        # Style data rows
        ExcelExportService.style_data_rows(ws, 4, row_num - 1, len(headers))
        
        return wb
    
    @staticmethod
    def export_factory_expenses():
        """Export Factory Expenses to Excel"""
        from models import FactoryExpense
        from app import db
        
        wb = ExcelExportService.create_styled_workbook()
        ws = wb.active
        ws.title = "Factory Expenses"
        
        # Add title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"Factory Expenses Export - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Headers
        headers = [
            "Expense Number", "Category", "Description", "Amount", 
            "Tax Amount", "Total Amount", "Status", "Date"
        ]
        
        ExcelExportService.style_header_row(ws, 3, headers)
        
        # Data
        expenses = db.session.query(FactoryExpense).all()
        row_num = 4
        
        for expense in expenses:
            data = [
                expense.expense_number, expense.category, 
                expense.description or '', expense.amount,
                expense.tax_amount or 0, expense.total_amount,
                expense.status, expense.expense_date.strftime('%Y-%m-%d')
            ]
            for col_num, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
            row_num += 1
        
        # Style data rows
        ExcelExportService.style_data_rows(ws, 4, row_num - 1, len(headers))
        
        return wb
    
    @staticmethod
    def create_response(workbook, filename):
        """Create Flask response for Excel download"""
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response